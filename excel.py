# excel.py
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from urllib.parse import quote
from config import RESULTS_FOLDER
from datetime import datetime

def read_resi_list(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File {file_path} tidak ditemukan!")
    if file_path.endswith('.csv'):
        df = pd.read_csv(file_path)
    else:
        df = pd.read_excel(file_path)
    resi_list = df["RESI"].dropna().astype(str).str.strip().tolist()
    return resi_list

def save_to_excel(data_list, output_path=None):
    if not data_list:
        print("Tidak ada data untuk disimpan.")
        return None

    # Buat folder results jika belum ada
    os.makedirs(RESULTS_FOLDER, exist_ok=True)

    # Generate nama file otomatis
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"hasil_followup_{timestamp}.xlsx"
        output_path = os.path.join(RESULTS_FOLDER, filename)

    df = pd.DataFrame(data_list)
    if "Pesan WA" not in df.columns:
        df["Pesan WA"] = ""

    # Urutkan kolom
    columns_order = ['Kode Resi', 'Status', 'Status Terakhir', 'Nama', 'HP', 'Paket', 'Nilai', 'Pesan WA']
    for col in columns_order:
        if col not in df.columns:
            df[col] = ""
    df = df[columns_order]

    df.to_excel(output_path, index=False, engine='openpyxl')
    
    # Tambahkan hyperlink WhatsApp
    wb = load_workbook(output_path)
    ws = wb.active

    hp_col = None
    pesan_col = None
    for col_idx, col_name in enumerate(ws[1], 1):
        if col_name.value == "HP":
            hp_col = col_idx
        elif col_name.value == "Pesan WA":
            pesan_col = col_idx

    if hp_col is not None and pesan_col is not None:
        link_col = pesan_col + 1
        ws.insert_cols(link_col)
        ws.cell(row=1, column=link_col, value="Link WA")

        for row in range(2, ws.max_row + 1):
            hp_cell = ws.cell(row=row, column=hp_col)
            pesan_cell = ws.cell(row=row, column=pesan_col)
            hp_value = hp_cell.value
            pesan_value = pesan_cell.value if pesan_cell.value else ""

            if hp_value and str(hp_value).strip():
                phone = ''.join(filter(str.isdigit, str(hp_value)))
                if phone:
                    if phone.startswith('0'):
                        phone = '62' + phone[1:]
                    encoded_msg = quote(pesan_value)
                    wa_url = f"https://wa.me/{phone}?text={encoded_msg}"
                    
                    link_cell = ws.cell(row=row, column=link_col)
                    link_cell.value = "Klik WhatsApp"
                    link_cell.hyperlink = wa_url
                    link_cell.font = Font(color="0000FF", underline="single")
                    link_cell.alignment = Alignment(horizontal='center')
        
        ws.column_dimensions[get_column_letter(link_col)].width = 18

    for col_letter in ['A','B','C','D','E','F','G','H','I']:
        ws.column_dimensions[col_letter].width = 25

    wb.save(output_path)
    print(f"\n[+] Berhasil menyimpan {len(data_list)} data ke {output_path}")
    return output_path